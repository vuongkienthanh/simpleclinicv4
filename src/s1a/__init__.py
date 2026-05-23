import argparse
from lib import DATE_FORMAT
from lib.db import connect
from lib.paths import DB, S1A_OUTPUT, S1A_TEMPLATE
from openpyxl import load_workbook
from openpyxl.styles import numbers
from copy import copy

# Create the parser
parser = argparse.ArgumentParser(
    prog="s1a", description="Tạo báo cáo s1a theo Thông tư 152/2025 "
)
# Add arguments
parser.add_argument("QUARTER", type=int, choices=[1, 2, 3, 4], help="quý")
parser.add_argument("YEAR", type=int, help="năm")


def main():
    # Parse the arguments
    args = parser.parse_args()
    if args.YEAR < 0:
        raise ValueError("YEAR should be positive")

    conn = connect(DB)
    result = conn.execute(
        """
        SELECT exam_datetime, (name || ' - ' || diagnosis) AS detail, price FROM visits
        JOIN patients
        WHERE visits.patient_id = patients.id 
            AND (
                CAST( strftime('%m', DATETIME(exam_datetime)) AS INTEGER) = ?
                OR CAST( strftime('%m', DATETIME(exam_datetime)) AS INTEGER) = ?
                OR CAST( strftime('%m', DATETIME(exam_datetime)) AS INTEGER) = ?
                )
            AND CAST( strftime('%Y', DATETIME(exam_datetime)) AS INTEGER) = ?
        """,
        (
            3 * (args.QUARTER - 1) + 1,
            3 * (args.QUARTER - 1) + 2,
            3 * (args.QUARTER - 1) + 3,
            args.YEAR,
        ),
    )
    wb = load_workbook(str(S1A_TEMPLATE))
    ws = wb.active
    assert ws is not None
    index = 9
    for row in result.fetchall():
        ws.insert_rows(index)
        ws[f"A{index}"] = row["exam_datetime"].Format(DATE_FORMAT)
        ws[f"B{index}"] = row["detail"]
        ws[f"C{index}"] = row["price"]
        ws[f"A{index}"].border = copy(ws[f"C{index+2}"].border)
        ws[f"B{index}"].border = copy(ws[f"C{index+2}"].border)
        ws[f"C{index}"].border = copy(ws[f"C{index+2}"].border)
        ws[f"C{index}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        index += 1
    wb.save(str(S1A_OUTPUT))
    wb.close()
